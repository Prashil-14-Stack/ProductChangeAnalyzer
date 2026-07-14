"""
==========================================================
Excel Formatting Helpers

Purpose
-------
Reusable formatting utilities for all Excel reports.

Responsibilities
----------------
✓ Titles
✓ Headers
✓ Body formatting
✓ Status colours
✓ Impact colours
✓ Confidence colours
✓ Zebra striping
✓ Auto-fit
✓ Freeze panes
✓ Filters

==========================================================
"""

from openpyxl.styles import Alignment

from openpyxl.utils import get_column_letter

from v2.reports.styles import *


class ExcelFormatter:

    # ======================================================
    # Title
    # ======================================================

    def apply_title(self, cell):

        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL
        cell.alignment = CENTER
        cell.border = MEDIUM_BORDER

    # ======================================================
    # Header
    # ======================================================

    def apply_header(self, cell):

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ======================================================
    # Sub Header
    # ======================================================

    def apply_sub_header(self, cell):

        cell.font = SUB_HEADER_FONT
        cell.fill = SUMMARY_FILL
        cell.alignment = LEFT
        cell.border = THIN_BORDER

    # ======================================================
    # Body
    # ======================================================

    def apply_body(self, cell):

        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER

    # ======================================================
    # Wrap Alignment
    # ======================================================

    def wrap_alignment(self):

        return Alignment(

            horizontal="left",

            vertical="top",

            wrap_text=True

        )

    # ======================================================
    # Zebra Striping
    # ======================================================

    def apply_zebra_row(

        self,

        worksheet,

        row

    ):

        if row % 2 == 0:

            fill = ALTERNATE_FILL

        else:

            fill = WHITE_FILL

        for cell in worksheet[row]:

            if cell.fill == HEADER_FILL:

                continue

            cell.fill = fill

    # ======================================================
    # Status Colours
    # ======================================================

    def apply_status_colour(

        self,

        cell,

        status

    ):

        if not status:

            return

        status = status.lower()

        if status == "added":

            cell.fill = ADDED_FILL

        elif status == "removed":

            cell.fill = REMOVED_FILL

        elif status == "modified":

            cell.fill = MODIFIED_FILL

        elif status == "unchanged":

            cell.fill = UNCHANGED_FILL

    # ======================================================
    # Impact Colours
    # ======================================================

    def apply_impact_colour(

        self,

        cell,

        impact

    ):

        if not impact:

            return

        impact = impact.lower()

        if impact == "high":

            cell.fill = HIGH_IMPACT_FILL

        elif impact == "medium":

            cell.fill = MEDIUM_IMPACT_FILL

        elif impact == "low":

            cell.fill = LOW_IMPACT_FILL

        elif impact == "none":

            cell.fill = NO_IMPACT_FILL

    # ======================================================
    # Confidence Colours
    # ======================================================

    def apply_confidence_colour(

        self,

        cell,

        confidence

    ):

        try:

            confidence = float(confidence)

        except Exception:

            return

        if confidence >= 90:

            cell.fill = HIGH_CONFIDENCE_FILL

        elif confidence >= 70:

            cell.fill = MEDIUM_CONFIDENCE_FILL

        else:

            cell.fill = LOW_CONFIDENCE_FILL

    # ======================================================
    # Auto Fit
    # ======================================================

    def autofit_columns(

        self,

        worksheet

    ):

        fixed_widths = {

            "A":18,
            "B":18,
            "C":30,
            "D":30,
            "E":16,
            "F":16,
            "G":18,
            "H":18,
            "I":18,
            "J":15,
            "K":45,
            "L":45,
            "M":40,
            "N":20,
            "O":15,
            "P":35,
            "Q":40,
            "R":35,
            "S":28,
            "T":25,
            "U":20,
            "V":18,
            "W":22

        }

        for column in worksheet.columns:

            letter = get_column_letter(

                column[0].column

            )

            if letter in fixed_widths:

                worksheet.column_dimensions[
                    letter
                ].width = fixed_widths[
                    letter
                ]

                continue

            max_length = 0

            for cell in column:

                if cell.value is None:

                    continue

                max_length = max(

                    max_length,

                    len(str(cell.value))

                )

            worksheet.column_dimensions[
                letter
            ].width = min(

                max_length + 3,

                60

            )

    # ======================================================
    # Freeze Header
    # ======================================================

    def freeze_header(

        self,

        worksheet,

        row=3

    ):

        worksheet.freeze_panes = f"A{row}"

    # ======================================================
    # Filter
    # ======================================================

    def apply_filter(
        self,
        worksheet,
        header_row=2
    ):
        """
        Apply filter only to the table header row and data,
        not the merged title row.
        """

        last_column = get_column_letter(worksheet.max_column)

        worksheet.auto_filter.ref = (
            f"A{header_row}:{last_column}{worksheet.max_row}"
        )

    # ======================================================
    # Row Heights
    # ======================================================

    def apply_default_row_heights(

        self,

        worksheet

    ):

        worksheet.row_dimensions[1].height = TITLE_ROW_HEIGHT

        worksheet.row_dimensions[2].height = HEADER_ROW_HEIGHT

        for row in range(

            3,

            worksheet.max_row + 1

        ):

            worksheet.row_dimensions[
                row
            ].height = BODY_ROW_HEIGHT