"""
==========================================================
Metrics Sheet

Purpose
-------
Creates KPI metrics and Excel charts.

Responsibilities
----------------
✓ KPI Summary
✓ Status Distribution
✓ Impact Distribution
✓ Excel Charts

==========================================================
"""

from openpyxl.chart import PieChart, Reference, BarChart
from v2.reports.formatting import ExcelFormatter


class MetricsSheet:

    # ======================================================
    # Constructor
    # ======================================================

    def __init__(self):

        self.formatter = ExcelFormatter()

    # ======================================================
    # Build
    # ======================================================

    def build(
        self,
        workbook,
        comparison_result
    ):

        ws = workbook.create_sheet(
            "Metrics"
        )

        # --------------------------------------------------
        # Title
        # --------------------------------------------------

        ws.merge_cells("A1:D1")

        ws["A1"] = "PRODUCT COMPARISON METRICS"

        self.formatter.apply_title(
            ws["A1"]
        )

        # --------------------------------------------------
        # Status Metrics
        # --------------------------------------------------

        row = 3

        ws["A3"] = "Status"

        ws["B3"] = "Count"

        self.formatter.apply_header(ws["A3"])
        self.formatter.apply_header(ws["B3"])

        summary = comparison_result.summary

        metrics = [

            ("Modified", summary.get("modified", 0)),

            ("Added", summary.get("added", 0)),

            ("Removed", summary.get("removed", 0)),

            ("Unchanged", summary.get("unchanged", 0))

        ]

        for status, count in metrics:

            ws[f"A{row+1}"] = status

            ws[f"B{row+1}"] = count

            self.formatter.apply_body(
                ws[f"A{row+1}"]
            )

            self.formatter.apply_body(
                ws[f"B{row+1}"]
            )

            row += 1

        # --------------------------------------------------
        # Impact Metrics
        # --------------------------------------------------

        row = 3

        ws["D3"] = "Impact"

        ws["E3"] = "Count"

        self.formatter.apply_header(ws["D3"])
        self.formatter.apply_header(ws["E3"])

        impacts = [

            ("High", summary.get("high_impact", 0)),

            ("Medium", summary.get("medium_impact", 0)),

            ("Low", summary.get("low_impact", 0)),

            ("None", summary.get("no_impact", 0))

        ]

        for impact, count in impacts:

            ws[f"D{row+1}"] = impact

            ws[f"E{row+1}"] = count

            self.formatter.apply_body(
                ws[f"D{row+1}"]
            )

            self.formatter.apply_body(
                ws[f"E{row+1}"]
            )

            row += 1

        # ==================================================
        # Pie Chart
        # ==================================================

        pie = PieChart()

        labels = Reference(

            ws,

            min_col=1,

            min_row=4,

            max_row=7

        )

        data = Reference(

            ws,

            min_col=2,

            min_row=3,

            max_row=7

        )

        pie.add_data(

            data,

            titles_from_data=True

        )

        pie.set_categories(labels)

        pie.title = "Parameter Status"

        ws.add_chart(

            pie,

            "A10"

        )

        # ==================================================
        # Bar Chart
        # ==================================================

        bar = BarChart()

        labels = Reference(

            ws,

            min_col=4,

            min_row=4,

            max_row=7

        )

        data = Reference(

            ws,

            min_col=5,

            min_row=3,

            max_row=7

        )

        bar.add_data(

            data,

            titles_from_data=True

        )

        bar.set_categories(labels)

        bar.title = "Impact Distribution"

        bar.y_axis.title = "Count"

        bar.x_axis.title = "Impact"

        ws.add_chart(

            bar,

            "J10"

        )

        # --------------------------------------------------
        # Formatting
        # --------------------------------------------------

        self.formatter.autofit_columns(
            ws
        )

        self.formatter.apply_default_row_heights(
            ws
        )