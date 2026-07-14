"""
==========================================================
Excel Styles

Purpose
-------
Central location for all Excel formatting.

Enterprise Theme
----------------
Professional blue corporate theme suitable for
Business Analysis reports.

==========================================================
"""

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

# ==========================================================
# Fonts
# ==========================================================

TITLE_FONT = Font(
    name="Calibri",
    size=18,
    bold=True,
    color="FFFFFF"
)

HEADER_FONT = Font(
    name="Calibri",
    size=11,
    bold=True,
    color="FFFFFF"
)

SUB_HEADER_FONT = Font(
    name="Calibri",
    size=11,
    bold=True,
    color="1F1F1F"
)

BODY_FONT = Font(
    name="Calibri",
    size=11,
    color="000000"
)

BODY_BOLD_FONT = Font(
    name="Calibri",
    size=11,
    bold=True,
    color="000000"
)

# ==========================================================
# Corporate Colours
# ==========================================================

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="003366"          # Dark Navy
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="005BAC"          # Corporate Blue
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)

WHITE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFFF"
)

ALTERNATE_FILL = PatternFill(
    fill_type="solid",
    fgColor="F7F9FC"
)

# ==========================================================
# Status Colours
# ==========================================================

ADDED_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3"
)

REMOVED_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC"
)

MODIFIED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

UNCHANGED_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9"
)

# ==========================================================
# Impact Colours
# ==========================================================

HIGH_IMPACT_FILL = PatternFill(
    fill_type="solid",
    fgColor="EA9999"
)

MEDIUM_IMPACT_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

LOW_IMPACT_FILL = PatternFill(
    fill_type="solid",
    fgColor="B6D7A8"
)

NO_IMPACT_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3"
)

# ==========================================================
# Confidence Colours
# ==========================================================

HIGH_CONFIDENCE_FILL = PatternFill(
    fill_type="solid",
    fgColor="B6D7A8"
)

MEDIUM_CONFIDENCE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFE599"
)

LOW_CONFIDENCE_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC"
)

# ==========================================================
# Borders
# ==========================================================

THIN = Side(
    style="thin",
    color="D0D7DE"
)

MEDIUM = Side(
    style="medium",
    color="7F8C8D"
)

THIN_BORDER = Border(
    left=THIN,
    right=THIN,
    top=THIN,
    bottom=THIN
)

MEDIUM_BORDER = Border(
    left=MEDIUM,
    right=MEDIUM,
    top=MEDIUM,
    bottom=MEDIUM
)

# ==========================================================
# Alignment
# ==========================================================

CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

LEFT = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

RIGHT = Alignment(
    horizontal="right",
    vertical="center",
    wrap_text=True
)

# ==========================================================
# Row Heights
# ==========================================================

TITLE_ROW_HEIGHT = 30

HEADER_ROW_HEIGHT = 24

BODY_ROW_HEIGHT = 34

# ==========================================================
# Default Column Widths
# ==========================================================

COLUMN_WIDTH_SMALL = 15

COLUMN_WIDTH_MEDIUM = 25

COLUMN_WIDTH_LARGE = 40

COLUMN_WIDTH_EXTRA_LARGE = 60