from openpyxl import Workbook

from openpyxl.styles import (

    Font,

    PatternFill,

    Border,

    Side,

    Alignment

)

from openpyxl.utils import get_column_letter

from openpyxl.formatting.rule import CellIsRule

from openpyxl.styles.differential import DifferentialStyle


class ExcelGeneratorDOCX:

    """
    Enterprise Excel Report Generator

    Generates a professional Product Change
    Analysis report for DOCX Product Specifications.

    Report Features

    • Enterprise Formatting
    • Auto Filters
    • Freeze Panes
    • Auto Column Width
    • Wrapped Text
    • Conditional Formatting
    • AI Business Intelligence
    """

    # ======================================================
    # Constructor
    # ======================================================

    def __init__(self):

        self.header_fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        self.header_font = Font(

            bold=True,

            color="FFFFFF",

            size=11

        )

        self.normal_font = Font(

            size=10

        )

        self.bold_font = Font(

            bold=True

        )

        self.wrap_alignment = Alignment(

            wrap_text=True,

            vertical="top"

        )

        self.center_alignment = Alignment(

            horizontal="center",

            vertical="center",

            wrap_text=True

        )

        thin = Side(

            style="thin",

            color="D9D9D9"

        )

        self.border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )

    # ======================================================
    # Public Method
    # ======================================================

    def generate_report(

        self,

        documents,

        comparison_table,

        analysis_results,

        file_name

    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Parameter Comparison"

        self._write_title(

            worksheet

        )

        self._write_headers(

            worksheet

        )

        self._write_rows(

            worksheet,

            comparison_table

        )

        self._format_sheet(

            worksheet

        )

        self._apply_conditional_formatting(

            worksheet,

            len(comparison_table)

        )

        self.finalize(

            worksheet,

            comparison_table

        )

        workbook.save(

            file_name

        )
    # ======================================================
    # Report Title
    # ======================================================

    def _write_title(

        self,

        worksheet

    ):

        worksheet.merge_cells("A1:V1")

        cell = worksheet["A1"]

        cell.value = "Product Change Analysis Report"

        cell.font = Font(

            bold=True,

            size=16,

            color="FFFFFF"

        )

        cell.fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )

        worksheet.row_dimensions[1].height = 28


    # ======================================================
    # Column Headers
    # ======================================================

    def _write_headers(

        self,

        worksheet

    ):

        headers = [

            "Source Version",

            "Target Version",

            "V1 Parameter",

            "Matched V2 Parameter",

            "Parameter Confidence",

            "Confidence Band",

            "Description Confidence",

            "Overall Confidence",

            "Decision",

            "Status",

            "V1 Content",

            "V2 Content",

            "Difference",

            "Change Type",

            "Severity",

            "Remarks",

            "Summary",

            "Business Impact",

            "Affected Teams",

            "Testing",

            "Risk",

            "Priority",

            "Business Criticality"

        ]

        row = 2

        for column, header in enumerate(headers, start=1):

            cell = worksheet.cell(

                row=row,

                column=column

            )

            cell.value = header

            cell.font = self.header_font

            cell.fill = self.header_fill

            cell.border = self.border

            cell.alignment = self.center_alignment
        
    # ======================================================
    # Write Report Rows
    # ======================================================

    def _write_rows(

        self,

        worksheet,

        comparison_table

    ):

        start_row = 3

        for row_number, row in enumerate(

            comparison_table,

            start=start_row

        ):

            values = [

                row.get("Source Version", ""),

                row.get("Target Version", ""),

                row.get("V1 Parameter", ""),

                row.get("Matched V2 Parameter", ""),

                row.get("Parameter Confidence", ""),

                row.get("Confidence Band", ""),

                row.get("Description Confidence", ""),

                row.get("Overall Confidence", ""),

                row.get("Decision", ""),

                row.get("Status", ""),

                row.get("V1", ""),

                row.get("V2", ""),

                row.get("Difference", ""),

                row.get("Change Type", ""),

                row.get("Severity", ""),

                row.get("Remarks", ""),

                row.get("Summary", ""),

                row.get("Business Impact", ""),

                row.get("Affected Teams", ""),

                row.get("Testing", ""),

                row.get("Risk", ""),

                row.get("Priority", ""),

                row.get("Business Criticality", "")

            ]

            for column_number, value in enumerate(

                values,

                start=1

            ):

                cell = worksheet.cell(

                    row=row_number,

                    column=column_number

                )

                cell.value = value

                cell.font = self.normal_font

                cell.border = self.border

                # -----------------------------------------
                # Wrap long business text
                # -----------------------------------------

                if column_number >= 11:

                    cell.alignment = self.wrap_alignment

                else:

                    cell.alignment = self.center_alignment

            # ---------------------------------------------
            # Increase row height automatically
            # ---------------------------------------------

            worksheet.row_dimensions[row_number].height = 55

        # ======================================================
    # Worksheet Formatting
    # ======================================================

    def _format_sheet(

        self,

        worksheet

    ):

        # ------------------------------------------
        # Freeze Header
        # ------------------------------------------

        worksheet.freeze_panes = "A3"

        # ------------------------------------------
        # Auto Filter
        # ------------------------------------------

        worksheet.auto_filter.ref = worksheet.dimensions

        # ------------------------------------------
        # Default Width
        # ------------------------------------------

        default_width = 18

        for column in worksheet.columns:

            column_letter = get_column_letter(

                column[0].column

            )

            worksheet.column_dimensions[

                column_letter

            ].width = default_width

        # ------------------------------------------
        # Wider Business Columns
        # ------------------------------------------

        custom_width = {

            "K": 45,   # V1 Content

            "L": 45,   # V2 Content

            "M": 50,   # Difference

            "P": 40,   # Remarks

            "Q": 40,   # Summary

            "R": 50,   # Business Impact

            "S": 35,   # Teams

            "T": 45,   # Testing

            "U": 18,   # Risk

            "V": 18,   # Priority

            "W": 20    # Business Criticality

        }

        for column, width in custom_width.items():

            worksheet.column_dimensions[

                column

            ].width = width

        # ------------------------------------------
        # Header Height
        # ------------------------------------------

        worksheet.row_dimensions[2].height = 25

        # ------------------------------------------
        # Center Header
        # ------------------------------------------

        for cell in worksheet[2]:

            cell.alignment = self.center_alignment

        # ------------------------------------------
        # Wrap all business text
        # ------------------------------------------

        for row in worksheet.iter_rows(

            min_row=3

        ):

            for cell in row:

                if cell.column >= 11:

                    cell.alignment = self.wrap_alignment

        # ======================================================
    # Conditional Formatting
    # ======================================================

    def _apply_conditional_formatting(

        self,

        worksheet,

        total_rows

    ):

        if total_rows == 0:

            return

        last_row = total_rows + 2

        # ==================================================
        # STATUS (Column J)
        # ==================================================

        green_fill = PatternFill(

            fill_type="solid",

            fgColor="C6EFCE"

        )

        red_fill = PatternFill(

            fill_type="solid",

            fgColor="FFC7CE"

        )

        orange_fill = PatternFill(

            fill_type="solid",

            fgColor="FCE4D6"

        )

        green_font = Font(

            color="006100",

            bold=True

        )

        red_font = Font(

            color="9C0006",

            bold=True

        )

        orange_font = Font(

            color="9C6500",

            bold=True

        )

        # ------------------------------------------
        # Status Colours
        # ------------------------------------------

        for row in range(3, last_row + 1):

            status = worksheet[f"J{row}"].value

            severity = worksheet[f"O{row}"].value

            confidence = worksheet[f"H{row}"].value

            # --------------------------------------
            # STATUS
            # --------------------------------------

            if status == "No Change":

                worksheet[f"J{row}"].fill = green_fill

                worksheet[f"J{row}"].font = green_font

            elif status == "Modified":

                worksheet[f"J{row}"].fill = orange_fill

                worksheet[f"J{row}"].font = orange_font

            elif status in (

                "No Match",

                "Added / Removed"

            ):

                worksheet[f"J{row}"].fill = red_fill

                worksheet[f"J{row}"].font = red_font

            # --------------------------------------
            # SEVERITY
            # --------------------------------------

            if severity == "High":

                worksheet[f"O{row}"].fill = red_fill

                worksheet[f"O{row}"].font = red_font

            elif severity == "Medium":

                worksheet[f"O{row}"].fill = orange_fill

                worksheet[f"O{row}"].font = orange_font

            elif severity == "Low":

                worksheet[f"O{row}"].fill = green_fill

                worksheet[f"O{row}"].font = green_font

            # --------------------------------------
            # OVERALL CONFIDENCE
            # --------------------------------------

            try:

                confidence = float(confidence)

            except Exception:

                continue

            if confidence >= 90:

                worksheet[f"H{row}"].fill = green_fill

                worksheet[f"H{row}"].font = green_font

            elif confidence >= 70:

                worksheet[f"H{row}"].fill = orange_fill

                worksheet[f"H{row}"].font = orange_font

            else:

                worksheet[f"H{row}"].fill = red_fill

                worksheet[f"H{row}"].font = red_font

        # ==================================================
        # Alternate Row Shading
        # ==================================================

        alternate_fill = PatternFill(

            fill_type="solid",

            fgColor="F8F9FA"

        )

        for row in range(3, last_row + 1):

            if row % 2 == 0:

                for cell in worksheet[row]:

                    if cell.fill.fill_type is None:

                        cell.fill = alternate_fill

        # ======================================================
    # Report Statistics
    # ======================================================

    def add_report_statistics(

        self,

        worksheet,

        comparison_table

    ):

        total = len(comparison_table)

        no_change = sum(

            1

            for row in comparison_table

            if row.get("Status") == "No Change"

        )

        modified = sum(

            1

            for row in comparison_table

            if row.get("Status") == "Modified"

        )

        unmatched = sum(

            1

            for row in comparison_table

            if row.get("Status") in (

                "Added / Removed",

                "No Match"

            )

        )

        start = max(

            len(comparison_table) + 5,

            8

        )

        worksheet[f"A{start}"] = "Report Summary"

        worksheet[f"A{start}"].font = Font(

            bold=True,

            size=13

        )

        worksheet[f"A{start+1}"] = "Total Parameters"

        worksheet[f"B{start+1}"] = total

        worksheet[f"A{start+2}"] = "No Change"

        worksheet[f"B{start+2}"] = no_change

        worksheet[f"A{start+3}"] = "Modified"

        worksheet[f"B{start+3}"] = modified

        worksheet[f"A{start+4}"] = "Added / Removed"

        worksheet[f"B{start+4}"] = unmatched

        for row in range(start + 1, start + 5):

            worksheet[f"A{row}"].font = self.bold_font

            worksheet[f"A{row}"].border = self.border

            worksheet[f"B{row}"].border = self.border

            worksheet[f"A{row}"].fill = PatternFill(

                fill_type="solid",

                fgColor="D9EAD3"

            )

    # ======================================================
    # Footer
    # ======================================================

    def add_footer(

        self,

        worksheet,

        comparison_table

    ):

        footer_row = max(

            len(comparison_table) + 12,

            15

        )

        worksheet.merge_cells(

            f"A{footer_row}:W{footer_row}"

        )

        cell = worksheet[f"A{footer_row}"]

        cell.value = (

            "Generated by Product Change Analyzer | "

            "Enterprise AI Business Impact Engine"

        )

        cell.font = Font(

            italic=True,

            color="666666",

            size=10

        )

        cell.alignment = Alignment(

            horizontal="center"

        )

    # ======================================================
    # Final Formatting
    # ======================================================

    def finalize(

        self,

        worksheet,

        comparison_table

    ):

        self.add_report_statistics(

            worksheet,

            comparison_table

        )

        self.add_footer(

            worksheet,

            comparison_table

        )