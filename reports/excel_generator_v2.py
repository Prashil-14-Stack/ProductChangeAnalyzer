from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ExcelGeneratorV2:

    def generate_report(

            self,

            documents,

            comparison_table,

            analysis_results,

            file_name

    ):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Parameter Comparison"

        # -----------------------------
        # Dynamic Header
        # -----------------------------

        headers = [

            "Source Version",

            "Target Version",

            "V1 Parameter",

            "Matched V2 Parameter",

            "Parameter Confidence",

            "Description Confidence",

            "Overall Confidence",

            "Decision",

            "Status",

            "Change Type",

            "Severity",

            "Remarks",

            "V1 Content",

            "V2 Content"

        ]


        sheet.append(headers)

        # -----------------------------
        # Header Styling
        # -----------------------------

        blue_fill = PatternFill(

            fill_type="solid",

            start_color="1F4E78",

            end_color="1F4E78"

        )

        white_font = Font(

            bold=True,

            color="FFFFFF"

        )

        for cell in sheet[1]:

            cell.fill = blue_fill

            cell.font = white_font

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center",

                wrap_text=True

            )

        # -----------------------------
        # Data
        # -----------------------------

        for row in comparison_table:

            sheet.append([

                row["Source Version"],

                row["Target Version"],

                row["V1 Parameter"],

                row["Matched V2 Parameter"],

                row["Parameter Confidence"],

                row["Description Confidence"],

                row["Overall Confidence"],

                row["Decision"],

                row["Status"],

                row["Change Type"],

                row["Severity"],

                row["Remarks"],

                row["V1"],

                row["V2"]

            ])

        # -----------------------------
        # Auto Width
        # -----------------------------

        for column in sheet.columns:

            width = 20

            letter = column[0].column_letter

            sheet.column_dimensions[letter].width = width

        workbook.save(file_name)