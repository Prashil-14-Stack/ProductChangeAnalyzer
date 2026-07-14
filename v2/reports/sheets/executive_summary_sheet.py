"""
==========================================================
Executive Summary Sheet

Purpose
-------
Displays the AI-generated executive summary in a
professional, readable format.

Responsibilities
----------------
✓ Executive Summary
✓ Key Business Changes
✓ Business Impact
✓ Recommendations

==========================================================
"""

from openpyxl.styles import Alignment

from v2.reports.formatting import ExcelFormatter


class ExecutiveSummarySheet:

    # ======================================================
    # Constructor
    # ======================================================

    def __init__(self):

        self.formatter = ExcelFormatter()

    # ======================================================
    # Build
    # ======================================================

    def build(
        self,
        workbook,
        comparison_result,
        business_summary=None
    ):

        ws = workbook.create_sheet(
            "Executive Summary"
        )

        # --------------------------------------------------
        # Title
        # --------------------------------------------------

        ws.merge_cells("A1:F1")

        ws["A1"] = "EXECUTIVE SUMMARY"

        self.formatter.apply_title(
            ws["A1"]
        )

        # --------------------------------------------------
        # Product Information
        # --------------------------------------------------

        row = 3

        product_info = [

            ("Product Version 1", comparison_result.product_name_v1),

            ("Product Version 2", comparison_result.product_name_v2),

            ("Version 1", comparison_result.version_v1),

            ("Version 2", comparison_result.version_v2)

        ]

        for label, value in product_info:

            ws[f"A{row}"] = label
            ws[f"B{row}"] = "" if value is None else str(value)

            self.formatter.apply_sub_header(
                ws[f"A{row}"]
            )

            self.formatter.apply_body(
                ws[f"B{row}"]
            )

            row += 1

        row += 2

        # --------------------------------------------------
        # AI Executive Summary
        # --------------------------------------------------

        ws[f"A{row}"] = "AI Executive Summary"

        self.formatter.apply_sub_header(
            ws[f"A{row}"]
        )

        row += 1

        if business_summary:

            sections = self._split_sections(
                business_summary
            )

            for heading, content in sections:

                # Section Heading

                ws.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=6
                )

                heading_cell = ws.cell(
                    row=row,
                    column=1
                )

                heading_cell.value = heading

                self.formatter.apply_sub_header(
                    heading_cell
                )

                row += 1

                # Section Content

                lines = max(
                    3,
                    content.count("\n") + 2
                )

                ws.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row + lines,
                    end_column=6
                )

                content_cell = ws.cell(
                    row=row,
                    column=1
                )

                content = self._escape_excel(content)
                content_cell.value = content

                content_cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

                self.formatter.apply_body(
                    content_cell
                )

                row += lines + 2

        else:

            ws.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row + 3,
                end_column=6
            )

            cell = ws.cell(
                row=row,
                column=1
            )

            cell.value = self._escape_excel(
                "No AI Business Summary Available."
            )

            self.formatter.apply_body(
                cell
            )

        # --------------------------------------------------
        # Formatting
        # --------------------------------------------------

        self.formatter.autofit_columns(
            ws
        )

        self.formatter.apply_default_row_heights(
            ws
        )

    # ======================================================
    # Helper
    # ======================================================

    def _split_sections(
        self,
        summary
    ):

        """
        Splits the AI summary into logical sections.

        Expected headings:

        EXECUTIVE SUMMARY

        KEY BUSINESS CHANGES

        BUSINESS IMPACT

        RECOMMENDED ACTIONS
        """

        headings = [

            "EXECUTIVE SUMMARY",

            "KEY BUSINESS CHANGES",

            "BUSINESS IMPACT",

            "RECOMMENDED ACTIONS"

        ]

        sections = []

        current_heading = "Summary"

        buffer = []

        for line in summary.splitlines():

            clean = line.strip()

            if clean.upper() in headings:

                if buffer:

                    sections.append(

                        (

                            current_heading,

                            "\n".join(buffer).strip()

                        )

                    )

                    buffer = []

                current_heading = clean.title()

            else:

                buffer.append(line)

        if buffer:

            sections.append(

                (

                    current_heading,

                    "\n".join(buffer).strip()

                )

            )

        return sections
    
    def _escape_excel(self, text):

        if text is None:
            return ""

        text = str(text)

        if text.startswith("="):
            text = "'" + text

        return text